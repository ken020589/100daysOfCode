##!/usr/bin/env python
#day10 openpyxl

import openpyxl
from openpyxl import Workbook
workbook = openpyxl.Workbook()
sheet = workbook.worksheets[0]
sheet['A1'] = 'Hello'
sheet['B1'] = 'Hello'
listTitle=['name','Phone']
sheet.append(listTitle)
workbook.save('test.xlsx')

workbook = openpyxl.load_workbook('test.xlsx')
sheet = workbook.worksheets[0]
sheet['B2'] = 'Height'
workbook.save('test.xlsx')

print('row:', sheet.max_row)
print('column:', sheet.max_column)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row = i, column = j).value)